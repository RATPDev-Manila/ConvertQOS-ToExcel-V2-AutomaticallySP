import pdfplumber
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import glob

# === SETTINGS ===
input_folder = r"C:\Users\Guillaume Turillon\RATP Dev\PowerAutomateAnalysis - Analysis\QoS-PDF"
sharepoint_root = r"C:\Users\Guillaume Turillon\RATP Dev\PowerAutomateAnalysis - Analysis\2025"

# === Find all matching PDFs ===
pdf_files = glob.glob(os.path.join(input_folder, "Consolidated Train Performance_*.pdf"))

if not pdf_files:
    print("No PDF files found.")
    for f in os.listdir(input_folder):
        print(f)
    exit()

for input_pdf in pdf_files:
    all_tables = []

    # Extract date and tables
    with pdfplumber.open(input_pdf) as pdf:
        first_page_text = pdf.pages[0].extract_text()
        match = re.search(r"Date\s*:\s*(\d{4})-(\d{2})-(\d{2})", first_page_text)
        if match:
            year, month, day = match.groups()
        else:
            print(f"Date not found in {input_pdf}, skipping...")
            continue

        # Build output folder and ensure it exists
        output_folder = os.path.join(sharepoint_root, month, day)
        os.makedirs(output_folder, exist_ok=True)
        output_excel = os.path.join(output_folder, f"QOS_{day}_{month}_{year}.xlsx")

        # Extract tables from all pages
        for i, page in enumerate(pdf.pages):
            print(f"Processing page {i+1}/{len(pdf.pages)} of {os.path.basename(input_pdf)}")
            tables = page.extract_tables()
            for table in tables:
                df = pd.DataFrame(table[1:], columns=table[0])
                df = df.applymap(lambda x: x.strip().replace('\u200b', '') if isinstance(x, str) else x)
                for col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='ignore')
                all_tables.append(df)

    # Combine tables
    if not all_tables:
        print(f"No tables found in {input_pdf}, skipping...")
        continue

    final_df = pd.concat(all_tables, ignore_index=True)

    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "QOS Data"

    # Define styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    light_row_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    dark_row_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Write header
    for col_idx, col_name in enumerate(final_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Write rows
    for row_idx, row in enumerate(final_df.itertuples(index=False), start=2):
        fill = light_row_fill if row_idx % 2 == 0 else dark_row_fill
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = center_alignment

    # Freeze header and autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save Excel file
    wb.save(output_excel)
    print(f"Saved: {output_excel}")

    # Delete processed PDF
    try:
        os.remove(input_pdf)
        print(f"Deleted: {input_pdf}")
    except Exception as e:
        print(f"Failed to delete {input_pdf}: {e}")
